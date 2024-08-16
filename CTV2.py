import sys
import os
from PyQt6.QtWidgets import QApplication, QWidget, QVBoxLayout, QPushButton, QFileDialog, QLabel, QMessageBox, QRadioButton, QStackedLayout, QProgressBar
from PyQt6.QtCore import QThread, pyqtSignal, pyqtSlot
from PyQt6.QtGui import QIcon

def resource_path(relative_path):
    """ Get the absolute path to the resource, works for dev and for PyInstaller """
    try:
        base_path = sys._MEIPASS  # PyInstaller stores data files in a temporary folder _MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

class Worker(QThread):
    update_status = pyqtSignal(str, str)
    progress_changed = pyqtSignal(int)

    def __init__(self, input_file_path, output_file_path, conversion_func, parent=None):
        super().__init__(parent)
        self.input_file_path = input_file_path
        self.output_file_path = output_file_path
        self.conversion_func = conversion_func

    def run(self):
        from PyQt6.QtWidgets import QApplication
        
        QApplication.processEvents()
        conversion_needed, file_path_to_use = self.ensure_utf8()

        if file_path_to_use is None:
            self.update_status.emit('Failed to prepare the file for conversion.', 'red')
            QApplication.processEvents()
            self.msleep(600)
            return

        QApplication.processEvents()
        self.msleep(1600)  # Allow UI to update after UTF-8 check

        try:
            self.conversion_func(file_path_to_use, self.output_file_path)
            self.update_status.emit('Conversion completed successfully!', 'green')
            self.progress_changed.emit(100)
            QApplication.processEvents()
        except Exception as e:
            self.update_status.emit(f'Conversion error:\n{e}', 'red')
            QApplication.processEvents()

    def ensure_utf8(self):
        import chardet
        import tempfile
        import shutil

        try:
            if self.input_file_path.lower().endswith('.csv'):
                with open(self.input_file_path, 'rb') as file:
                    raw_data = file.read(4096)  # Small portion for encoding detection
                    detection = chardet.detect(raw_data)
                    encoding = detection.get('encoding', 'utf-8').lower()

                if encoding == 'utf-8' or encoding == 'ascii':
                    self.update_status.emit('The CSV is UTF-8. Proceeding with conversion...', 'green')
                    return False, self.input_file_path  # No conversion needed, use original path
                else:
                    self.update_status.emit(f'File is not UTF-8 ({encoding}).\nEncoding to UTF-8 before converting...', '#E59400')
                    temp_file = tempfile.NamedTemporaryFile(delete=False, mode='w+', encoding='utf-8', suffix='.csv', newline='')
                    with open(self.input_file_path, 'r', encoding=encoding) as source_file:
                        shutil.copyfileobj(source_file, temp_file)
                    temp_file_path = temp_file.name
                    temp_file.close()  # Close the file so it can be reopened later in another mode if necessary
                    return True, temp_file_path  # Conversion performed, use temporary file path
            elif self.input_file_path.lower().endswith(('.xlsx', '.xls')):
                self.update_status.emit('Checking Excel file encoding...', 'green')
                return False, self.input_file_path
        except Exception as e:
            self.update_status.emit(f'Error during encoding detection or conversion: {e}', 'red')
            return False, None

class StepWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent=parent)
        self.layout = QVBoxLayout(self)

    def nextStep(self):
        pass

    def previousStep(self):
        pass

class FileSelectionStep(StepWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.selectFileBtn = QPushButton('Select Input File')
        self.selectFileBtn.setToolTip("Supported formats: .csv, .xlsx, .xls")
        self.selectFileBtn.clicked.connect(self.openFileNameDialog)
        self.layout.addWidget(self.selectFileBtn)

        self.selectedFileLabel = QLabel('No File Selected')
        self.layout.addWidget(self.selectedFileLabel)

    def openFileNameDialog(self):
        file_filter = "All Supported Files (*.csv *.xlsx *.xls);;CSV Files (*.csv);;Excel Files (*.xlsx *.xls)"
        fileName, _ = QFileDialog.getOpenFileName(self, "Select File", "", file_filter)
        if fileName:
            basename = os.path.basename(fileName)
            self.selectedFileLabel.setText(basename)  # Display only the file name
            self.parent().inputFilePath = fileName
            self.selectedFileLabel.setToolTip(basename)  # Set tooltip to full basename
            self.selectedFileLabel.setMaximumWidth(220)  # Adjust the value as needed
            self.selectedFileLabel.setWordWrap(True)
            self.parent().updateButtons()

class ConversionTypeStep(StepWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.radioToTSV = QRadioButton("Convert to TSV")
        self.radioToXLSX = QRadioButton("Convert to XLSX (Excel)")
        self.radioToCSVFromExcel = QRadioButton("Convert Excel to CSV")

        self.layout.addWidget(self.radioToTSV)
        self.layout.addWidget(self.radioToXLSX)
        self.layout.addWidget(self.radioToCSVFromExcel)

        self.radioToTSV.setChecked(True)  # Default Selection

class OutputSelectionStep(StepWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.selectOutputBtn = QPushButton('Select Output File Location')
        self.selectOutputBtn.clicked.connect(self.openSaveFileDialog)
        self.layout.addWidget(self.selectOutputBtn)

        self.selectedOutputLabel = QLabel('No Output Location Selected')
        self.layout.addWidget(self.selectedOutputLabel)

    def openSaveFileDialog(self):
        outputDir = QFileDialog.getExistingDirectory(self, "Select Output Folder")
        if outputDir:
            input_basename = os.path.basename(self.parent().inputFilePath)
            output_basename, _ = os.path.splitext(input_basename)
            if self.parent().conversionTypeStep.radioToTSV.isChecked():
                output_file = output_basename + '.tsv'
            elif self.parent().conversionTypeStep.radioToXLSX.isChecked():
                output_file = output_basename + '.xlsx'
            elif self.parent().conversionTypeStep.radioToCSVFromExcel.isChecked():
                output_file = output_basename + '.csv'

            outputFilePath = os.path.join(outputDir, output_file)
            self.selectedOutputLabel.setText(output_file)  # Display only the file name
            self.parent().outputFilePath = outputFilePath  # Store the full path in the parent class
            self.selectedOutputLabel.setToolTip(outputFilePath)  # Set tooltip to full output path
            self.selectedOutputLabel.setMaximumWidth(220)
            self.selectedOutputLabel.setWordWrap(True)
            self.parent().updateButtons()
        else:
            QMessageBox.information(self, "No Folder Selected", "You did not select any folder.")

class ConfirmationStep(StepWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.confirmBtn = QPushButton('Convert')
        self.confirmBtn.clicked.connect(self.parent().convertFile)
        self.layout.addWidget(self.confirmBtn)

        self.progressBar = QProgressBar(self)
        self.progressBar.setRange(0, 100)
        self.layout.addWidget(self.progressBar)

class ConverterWizard(QWidget):
    def __init__(self):
        super().__init__()
        self.title = 'OCT Wizard'
        self.inputFilePath = ''
        self.outputFilePath = ''
        self.initUI()
        self.statusLabel = QLabel('', self)
        self.layout.addWidget(self.statusLabel)

    def initUI(self):
        self.setWindowTitle(self.title)
        self.centerWindow()
        self.setFixedSize(400, 300)

        self.layout = QVBoxLayout(self)

        self.steps = QStackedLayout()
        self.fileSelectionStep = FileSelectionStep(self)
        self.conversionTypeStep = ConversionTypeStep(self)
        self.outputSelectionStep = OutputSelectionStep(self)
        self.confirmationStep = ConfirmationStep(self)

        self.steps.addWidget(self.fileSelectionStep)
        self.steps.addWidget(self.conversionTypeStep)
        self.steps.addWidget(self.outputSelectionStep)
        self.steps.addWidget(self.confirmationStep)

        self.layout.addLayout(self.steps)

        self.nextButton = QPushButton("Next")
        self.nextButton.clicked.connect(self.nextStep)
        self.layout.addWidget(self.nextButton)

        self.backButton = QPushButton("Back")
        self.backButton.clicked.connect(self.previousStep)
        self.layout.addWidget(self.backButton)

        self.updateButtons()

    def nextStep(self):
        currentIndex = self.steps.currentIndex()
        if currentIndex < self.steps.count() - 1:
            self.steps.setCurrentIndex(currentIndex + 1)
        self.updateButtons()

    def previousStep(self):
        currentIndex = self.steps.currentIndex()
        if currentIndex > 0:
            self.steps.setCurrentIndex(currentIndex - 1)
        self.updateButtons()

    def updateButtons(self):
        currentIndex = self.steps.currentIndex()
        self.nextButton.setDisabled(currentIndex == self.steps.count() - 1)
        self.backButton.setDisabled(currentIndex == 0)
        if currentIndex == 0:
            self.nextButton.setDisabled(self.inputFilePath == '')
        elif currentIndex == 2:
            self.nextButton.setDisabled(self.outputFilePath == '')

    def centerWindow(self):
        screen = QApplication.primaryScreen().geometry()
        x = (screen.width() - self.width()) // 2
        y = (screen.height() - self.height()) // 2
        self.move(x, y)

    @pyqtSlot()
    def convertFile(self):
        input_file = self.inputFilePath
        output_file = self.outputFilePath

        if input_file == '' or output_file == '':
            self.updateStatusLabel("Error: Please select both an input file and an output file location.", "red")
            return

        if input_file.lower().endswith('.csv') and self.conversionTypeStep.radioToXLSX.isChecked():
            conversion_func = self.convert_csv_to_xlsx
        elif input_file.lower().endswith('.csv') and self.conversionTypeStep.radioToTSV.isChecked():
            conversion_func = self.convert_csv_to_tsv
        elif input_file.lower().endswith('.xlsx') or input_file.lower().endswith('.xls'):
            if self.conversionTypeStep.radioToCSVFromExcel.isChecked():
                conversion_func = self.convert_excel_to_csv
            else:
                self.updateStatusLabel('Unsupported file format or operation.', 'red')
                return
        else:
            self.updateStatusLabel('Unsupported file format or operation.', 'red')
            return

        self.worker = Worker(input_file, output_file, conversion_func, self)
        self.worker.update_status.connect(self.updateStatusLabel)
        self.worker.progress_changed.connect(self.confirmationStep.progressBar.setValue)
        self.worker.start()

    def updateStatusLabel(self, text, color):
        self.statusLabel.setText(text)
        self.statusLabel.setStyleSheet(f'color: {color};')
        QApplication.processEvents()  # Process events to update label immediately
        if 'successfully' in text:
            self.showSuccessPrompt()

    def showSuccessPrompt(self):
        reply = QMessageBox.question(self, 'Conversion Successful',
                                     'Your file has been converted. Would you like to convert something else?',
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            self.reset()
        else:
            QApplication.quit() # This method will ensure the application exits in a clean mannor

    def reset(self):
        self.inputFilePath = ''
        self.outputFilePath = ''
        self.steps.setCurrentIndex(0)
        self.fileSelectionStep.selectedFileLabel.setText('No File Selected')
        self.outputSelectionStep.selectedOutputLabel.setText('No Output Location Selected')
        self.statusLabel.setText('')
        self.confirmationStep.progressBar.setValue(0)
        self.updateButtons()

    def convert_csv_to_tsv(self, input_file_path, output_file_path):
        import csv
        with open(input_file_path, mode='r', newline='', encoding='utf-8') as csvfile, \
                open(output_file_path, mode='w', newline='', encoding='utf-8') as tsvfile:
            csv_reader = csv.reader(csvfile, delimiter=',')
            csv_writer = csv.writer(tsvfile, delimiter='\t')
            total_rows = sum(1 for _ in csv_reader)
            csvfile.seek(0)  # Reset the file reader to the beginning
            for row_index, row in enumerate(csv_reader, start=1):
                csv_writer.writerow(row)
                self.worker.progress_changed.emit(int((row_index / total_rows) * 100))

    def convert_csv_to_xlsx(self, input_file_path, output_file_path):
        import csv
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        
        with open(input_file_path, mode='r', newline='', encoding='utf-8') as csvfile:
            csv_reader = csv.reader(csvfile, delimiter=',')
            
            # First, count the total rows
            total_rows = sum(1 for _ in csv_reader)
            csvfile.seek(0)  # Reset the file reader to the beginning

            for row_index, row in enumerate(csv_reader, start=1):
                for column_index, cell_value in enumerate(row, start=1):
                    cell = ws.cell(row=row_index, column=column_index)
                    cell.value = cell_value
                    column_letter = get_column_letter(column_index)
                    ws.column_dimensions[column_letter].width = max(ws.column_dimensions[column_letter].width, len(cell_value))

                self.worker.progress_changed.emit(int((row_index / total_rows) * 100))
            
            wb.save(output_file_path)

    def convert_excel_to_csv(self, input_file_path, output_file_path):
        import openpyxl
        import csv
        try:
            wb = openpyxl.load_workbook(input_file_path)
            sh = wb.active
            total_rows = sh.max_row
            with open(output_file_path, 'w', newline='', encoding='utf-8') as f:
                c = csv.writer(f)
                for row_index, r in enumerate(sh.rows, start=1):
                    c.writerow([cell.value for cell in r])
                    self.worker.progress_changed.emit(int((row_index / total_rows) * 100))
            self.updateStatusLabel('Excel file converted to CSV successfully!', 'green')
        except Exception as e:
            self.updateStatusLabel(f'Excel to CSV conversion error: {e}', 'red')

if __name__ == '__main__':
    app = QApplication(sys.argv)
    
    # Set the global application icon before creating any windows
    icon_path = resource_path('OCT.png')
    print(f"Icon path: {icon_path}")
    app.setWindowIcon(QIcon(icon_path))
    
    wizard = ConverterWizard()

    # Use resource_path to find the correct path to style.qss
    style_file = resource_path('style.qss')
    with open(style_file, "r") as file:
        app.setStyleSheet(file.read())
    
    wizard.show()
    sys.exit(app.exec())